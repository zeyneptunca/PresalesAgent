import csv
import io


def _write_csv(filepath: str, headers: list[str], rows: list[list]) -> None:
    """UTF-8 BOM + ; ayrac ile CSV yazar."""
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)


def write_wp_details(result: dict, filepath: str) -> None:
    """Her satir 1 WP olan detay CSV'si."""
    headers = [
        "modul", "wp_id", "wp_adi", "complexity", "deliverable_sayisi",
        "baskin_kategori", "of_eslesmesi", "reuse_durumu",
        "a_analiz", "a_tasarim", "a_mimari", "a_fe", "a_be", "a_test", "a_toplam",
        "b_analiz", "b_tasarim", "b_mimari", "b_fe", "b_be", "b_test", "b_toplam",
        "c_analiz", "c_tasarim", "c_mimari", "c_fe", "c_be", "c_test", "c_toplam",
        "min_a", "max_a", "min_b", "max_b", "min_c", "max_c",
        "hesaplama_hikayesi",
    ]

    rows = []
    for wp in result.get("wp_detaylari", []):
        row = [wp.get(h, "") for h in headers]
        rows.append(row)

    _write_csv(filepath, headers, rows)


def write_summary(result: dict, filepath: str) -> None:
    """Dikey format ozet CSV'si."""
    headers = ["bolum", "kalem", "profil_a", "profil_b", "profil_c"]
    rows = []

    ozet = result.get("tahmin_ozeti", {})
    rows.append(["PROJE", "proje_adi", ozet.get("proje_adi", ""), "", ""])
    rows.append(["PROJE", "tarih", ozet.get("tahmin_tarihi", ""), "", ""])
    rows.append(["PROJE", "toplam_modul", ozet.get("toplam_modul", ""), "", ""])
    rows.append(["PROJE", "toplam_wp", ozet.get("toplam_wp", ""), "", ""])
    rows.append(["PROJE", "proje_bandi", "-",
                 ozet.get("proje_bandi", ""), ozet.get("proje_bandi", "")])

    for mt in result.get("modul_toplamlari", []):
        rows.append(["MODUL", mt.get("modul", ""),
                     mt.get("a_toplam", 0), mt.get("b_toplam", 0), mt.get("c_toplam", 0)])

    faz = result.get("faz_toplamlari", {})
    for profil_key in ["analiz", "tasarim", "mimari", "fe", "be", "test", "toplam"]:
        rows.append(["FAZ", profil_key,
                     faz.get("a", {}).get(profil_key, 0),
                     faz.get("b", {}).get(profil_key, 0),
                     faz.get("c", {}).get(profil_key, 0)])

    gl = result.get("global_eforlar", {})
    for g_key in ["pm", "tech_design", "devops", "deployment", "uat",
                   "ba_base", "test_base", "uat_base", "toplam"]:
        label = "global_toplam" if g_key == "toplam" else g_key
        rows.append(["GLOBAL", label,
                     gl.get("a", {}).get(g_key, 0),
                     gl.get("b", {}).get(g_key, 0),
                     gl.get("c", {}).get(g_key, 0)])

    pt = result.get("proje_toplami", {})
    for t_key in ["teknik", "global", "toplam", "min", "max"]:
        label = "genel_toplam" if t_key == "toplam" else t_key
        rows.append(["TOPLAM", label,
                     pt.get("a", {}).get(t_key, 0),
                     pt.get("b", {}).get(t_key, 0),
                     pt.get("c", {}).get(t_key, 0)])

    rows.append(["TOPLAM", "tasarruf_ag", "-",
                 pt.get("tasarruf_b_ag", 0), pt.get("tasarruf_c_ag", 0)])
    rows.append(["TOPLAM", "tasarruf_yuzde", "-",
                 pt.get("tasarruf_b_yuzde", 0), pt.get("tasarruf_c_yuzde", 0)])

    bc = result.get("baglam_carpanlari", {})
    rows.append(["BAGLAM", "faktor",
                 bc.get("faktor_a_b", 0), bc.get("faktor_a_b", 0), bc.get("faktor_c", 0)])

    _write_csv(filepath, headers, rows)


def write_notes(result: dict, filepath: str) -> None:
    """Not, risk ve kapsam disi CSV'si."""
    headers = ["tip", "icerik"]
    rows = []

    for note in result.get("notlar", []):
        if note and str(note).strip():
            rows.append(["NOT", note])

    for risk in result.get("riskler", []):
        if risk and str(risk).strip():
            rows.append(["RISK", risk])

    for kd in result.get("kapsam_disi", []):
        if kd and str(kd).strip():
            rows.append(["KAPSAM_DISI", kd])

    _write_csv(filepath, headers, rows)


def _join_list(items: list, sep: str = " | ") -> str:
    """Liste elemanlarini string olarak birlestirir."""
    if not items:
        return ""
    str_items = []
    for item in items:
        if isinstance(item, dict):
            str_items.append(item.get("name", item.get("adi", str(item))))
        else:
            str_items.append(str(item))
    return sep.join(str_items)


def write_wbs_details(wbs: dict, filepath: str) -> None:
    """WBS yapisi — her satir 1 WP, tum icerik alanlariyla birlikte."""
    headers = [
        "modul_id", "modul_adi", "modul_aciklama",
        "wp_id", "wp_adi", "wp_aciklama",
        "complexity", "complexity_drivers",
        "deliverables", "deliverable_sayisi",
        "frontend_requirements", "backend_requirements", "data_implications",
        "integration_points", "entegrasyon_sayisi",
        "acceptance_criteria",
    ]

    rows = []
    for mod in wbs.get("wbs", {}).get("modules", []):
        mod_id = mod.get("module_id", "")
        mod_name = mod.get("name", "")
        mod_desc = mod.get("description", "")

        for wp in mod.get("work_packages", []):
            tc = wp.get("technical_context", {})
            int_points = tc.get("integration_points", [])
            deliverables = wp.get("deliverables", [])
            drivers = wp.get("complexity", {}).get("drivers", [])
            criteria = wp.get("acceptance_criteria", [])

            rows.append([
                mod_id, mod_name, mod_desc,
                wp.get("wp_id", ""), wp.get("name", ""), wp.get("description", ""),
                wp.get("complexity", {}).get("level", ""),
                _join_list(drivers),
                _join_list(deliverables),
                len(deliverables),
                tc.get("frontend_requirements", ""),
                tc.get("backend_requirements", ""),
                tc.get("data_implications", ""),
                _join_list(int_points),
                len(int_points),
                _join_list(criteria),
            ])

    _write_csv(filepath, headers, rows)


def write_categorization_details(categories: dict, filepath: str) -> None:
    """Kategorizasyon detaylari — her satir 1 deliverable."""
    headers = [
        "wp_id", "baskin_kategori",
        "deliverable_adi", "kategori", "of_match", "neden",
    ]

    rows = []
    wp_cat = categories.get("wp_kategorileri", {})
    for wp_id, cat_data in wp_cat.items():
        baskin = cat_data.get("baskin_kategori", "")
        deliverables = cat_data.get("deliverables", [])
        for d in deliverables:
            rows.append([
                wp_id, baskin,
                d.get("adi", d.get("name", "")),
                d.get("kategori", ""),
                d.get("of_match", ""),
                d.get("neden", ""),
            ])

    _write_csv(filepath, headers, rows)


def write_context_analysis(categories: dict, effort_result: dict, filepath: str) -> None:
    """Baglam analizi ve carpanlar — dikey format."""
    headers = ["bolum", "kalem", "deger", "aciklama"]
    rows = []

    baglam = categories.get("baglam_analizi", {})

    # Domain
    rows.append(["DOMAIN", "karmasiklik",
                 baglam.get("domain_karmasikligi", ""),
                 baglam.get("domain_neden", "")])

    # Entegrasyon
    rows.append(["ENTEGRASYON", "yogunluk",
                 baglam.get("entegrasyon_yogunlugu", ""),
                 baglam.get("entegrasyon_neden", "")])

    ent_list = baglam.get("benzersiz_entegrasyonlar", [])
    if ent_list:
        rows.append(["ENTEGRASYON", "liste", ", ".join(ent_list),
                     f"{len(ent_list)} benzersiz entegrasyon"])

    # Reuse gruplari
    reuse = categories.get("reuse_gruplari", {})
    if isinstance(reuse, dict):
        for group_id, group_data in reuse.items():
            if isinstance(group_data, dict):
                wps = group_data.get("wp_ids", group_data.get("wps", []))
                neden = group_data.get("neden", group_data.get("reason", ""))
                rows.append(["REUSE", group_id,
                             ", ".join(wps) if isinstance(wps, list) else str(wps),
                             neden])
            elif isinstance(group_data, list):
                rows.append(["REUSE", group_id, ", ".join(str(x) for x in group_data), ""])
    elif isinstance(reuse, list):
        for i, group in enumerate(reuse):
            rows.append(["REUSE", f"grup_{i+1}", str(group), ""])

    # Baglam carpanlari
    bc = effort_result.get("baglam_carpanlari", {})
    carpan_labels = {
        "olcek": "Organizasyon Olcegi",
        "ekip": "Ekip Deneyimi",
        "domain": "Domain Karmasikligi",
        "teknik_borc": "Teknik Borc",
        "ent_yogunluk": "Entegrasyon Yogunlugu",
        "faktor_a_b": "Birlesik Faktor (A ve B)",
        "faktor_c": "Birlesik Faktor (C, max 1.20)",
    }
    for key, label in carpan_labels.items():
        val = bc.get(key, "")
        if val != "":
            if isinstance(val, dict):
                rows.append(["CARPAN", key, f"AB={val.get('AB', 1.0)} C={val.get('C', 1.0)}", label])
            else:
                rows.append(["CARPAN", key, val, label])

    _write_csv(filepath, headers, rows)


def write_full_export(wbs: dict, categories: dict, effort_result: dict,
                      filepath: str) -> None:
    """Tek Excel dosyasi (.xlsx) — 6 sheet ile tum veri.
    Sayisal degerler gercek sayi olarak yazilir, boylece Excel
    kullanicinin locale ayarina gore (TR: virgul) gosterir."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl kurulu degil. 'pip install openpyxl' ile kurun.")

    wb = openpyxl.Workbook()

    def _try_numeric(value):
        """String degeri mumkunse int veya float'a cevir."""
        if not isinstance(value, str):
            return value
        s = value.strip()
        if not s:
            return s
        # int dene
        try:
            return int(s)
        except ValueError:
            pass
        # float dene
        try:
            return float(s)
        except ValueError:
            pass
        return value

    def _csv_to_sheet(sheet, writer_fn, *args):
        """CSV writer fonksiyonunu cagirip sonucu sheet'e yazar.
        Sayisal hucreleri gercek sayi olarak yazar."""
        import tempfile as _tf
        import os
        with _tf.NamedTemporaryFile(delete=False, suffix=".csv", mode="w") as tmp:
            tmp_path = tmp.name
        writer_fn(*args, tmp_path)

        with open(tmp_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=";")
            first_row = True
            for row in reader:
                if first_row:
                    # Header satiri string kalsin
                    sheet.append(row)
                    first_row = False
                else:
                    # Veri satirlarinda sayisal degerleri cevir
                    sheet.append([_try_numeric(cell) for cell in row])
        os.unlink(tmp_path)

    # Sheet 1: WBS Yapisi
    ws1 = wb.active
    ws1.title = "WBS Yapisi"
    _csv_to_sheet(ws1, write_wbs_details, wbs)

    # Sheet 2: Kategorizasyon
    ws2 = wb.create_sheet("Kategorizasyon")
    _csv_to_sheet(ws2, write_categorization_details, categories)

    # Sheet 3: Efor Detaylari
    ws3 = wb.create_sheet("Efor Detaylari")
    _csv_to_sheet(ws3, write_wp_details, effort_result)

    # Sheet 4: Ozet
    ws4 = wb.create_sheet("Ozet")
    _csv_to_sheet(ws4, write_summary, effort_result)

    # Sheet 5: Baglam ve Carpanlar
    ws5 = wb.create_sheet("Baglam ve Carpanlar")
    _csv_to_sheet(ws5, write_context_analysis, categories, effort_result)

    # Sheet 6: Notlar ve Riskler
    ws6 = wb.create_sheet("Notlar ve Riskler")
    _csv_to_sheet(ws6, write_notes, effort_result)

    wb.save(filepath)
